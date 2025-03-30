#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
文件处理模块
负责导入和导出数据文件
"""

import os
import pandas as pd
import numpy as np


class FileHandler:
    """文件处理类，负责数据导入和结果导出"""
    
    def load_data(self, filename):
        """
        加载数据文件
        
        参数:
            filename (str): 文件路径
            
        返回:
            tuple: (numbers, indices) 数字列表和对应的原始索引
        """
        ext = os.path.splitext(filename)[1].lower()
        
        if ext in ['.xlsx', '.xls']:
            return self._load_excel(filename)
        elif ext == '.txt':
            return self._load_txt(filename)
        else:
            raise ValueError("不支持的文件格式")
    
    def _load_excel(self, filename):
        """加载Excel文件"""
        df = pd.read_excel(filename, header=None)
        
        # 考虑可能是单列或单行数据
        if df.shape[0] == 1:  # 单行
            numbers = df.iloc[0, :].dropna().tolist()
        else:  # 默认取第一列
            numbers = df.iloc[:, 0].dropna().tolist()
        
        # 验证数据
        self._validate_data(numbers)
        
        # 创建原始索引
        indices = list(range(len(numbers)))
        
        return numbers, indices
    
    def _load_txt(self, filename):
        """加载文本文件"""
        numbers = []
        with open(filename, 'r') as f:
            lines = f.readlines()
            for line in lines:
                try:
                    val = float(line.strip())
                    numbers.append(val)
                except ValueError:
                    pass  # 忽略无法转换的行
        
        # 验证数据
        self._validate_data(numbers)
        
        # 创建原始索引
        indices = list(range(len(numbers)))
        
        return numbers, indices
    
    def _validate_data(self, numbers):
        """验证数据有效性"""
        if not numbers:
            raise ValueError("文件中没有有效数据")
        
        if len(numbers) < 10 or len(numbers) > 200:
            raise ValueError(f"数据数量 ({len(numbers)}) 不在有效范围内 (10-200)")
        
        # 检查是否都是数字且最多两位小数
        for num in numbers:
            if not isinstance(num, (int, float)):
                raise ValueError("数据中包含非数字值")
            
            # 检查小数位数
            str_num = str(num)
            if '.' in str_num:
                decimal_places = len(str_num.split('.')[1])
                if decimal_places > 2:
                    raise ValueError(f"数据 {num} 超过两位小数")
    
    def export_results(self, filename, numbers, indices, results, target=None):
        """
        导出结果到Excel文件
        
        参数:
            filename (str): 导出文件路径
            numbers (list): 原始数字列表
            indices (list): 原始索引列表
            results (list): 计算结果，每个元素是一个包含选中索引的列表
            target (float, optional): 目标和值，用于验证计算
        """
        # 创建数据框
        df = pd.DataFrame()
        
        # 添加原始数据列
        df['数值'] = numbers
        
        # 为每个解添加一列
        for i, result in enumerate(results):
            column_name = f'解 {i+1}'
            df[column_name] = 0
            
            # 标记选中的数字
            for idx in result:
                original_idx = indices.index(idx)
                df.loc[original_idx, column_name] = 1
        
        # 添加是否被选中的列
        df['是否选中'] = df.iloc[:, 1:].any(axis=1).astype(int)
        
        # 计算和显示每个解的总和
        if results:
            # 添加总和行
            sums_row = len(df)
            df.loc[sums_row] = [None] * len(df.columns)
            df.loc[sums_row, '数值'] = '总和'
            
            for i, result in enumerate(results):
                column_name = f'解 {i+1}'
                # 计算总和（使用完整精度）
                result_sum = sum(numbers[indices.index(idx)] for idx in result)
                # 显示在总和行
                df.loc[sums_row, column_name] = result_sum
                
                # 如果提供了目标值，添加与目标的差值
                if target is not None:
                    diff = abs(result_sum - target)
                    df.loc[sums_row+1, '数值'] = '与目标差值'
                    df.loc[sums_row+1, column_name] = diff
        
        # 保存到Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='计算结果')
            
            # 设置条件格式（黄色背景）
            workbook = writer.book
            worksheet = writer.sheets['计算结果']
            
            # 使用openpyxl设置条件格式
            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # 为每行应用格式
            for row in range(2, len(df) + 2):  # Excel行从1开始，有表头所以+2
                if df.loc[row-2, '是否选中'] == 1:
                    for col in range(1, 3):  # 只对前两列应用黄色背景
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = yellow_fill
